"""
Database layer for the Hygiene Vending Machine.
Handles SQLite connection, schema, seeding, and all product/transaction/RFID/admin queries.
"""
import hashlib
import hmac as _hmac
import secrets
import sqlite3
from pathlib import Path
from datetime import datetime as dt, timedelta

from openpyxl import Workbook

from admin.reports import get_reports_dir


BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / "vending.db"

# Canonical machine stock capacities (source of truth)
_CAPACITY_RULES = {
    "alcohol": 3,
    "soap": 7,
    "deodorant": 9,
    "deo": 9,
    "mouthwash": 7,
    "mouth wash": 7,
    "wet wipes": 3,
    "wetwipes": 3,
    "wipes": 3,
    "tissue": 3,
    "tissues": 3,
    "panty liner": 8,
    "panty liners": 8,
    "pantyliners": 8,
    "panti liner": 8,
    "all night pads": 6,
    "all-night pads": 6,
    "regular with wings": 7,
    "regular w/ wings pads": 7,
    "regular with wings pads": 7,
    "non wings pad": 7,
    "non wing pad": 7,
    "non-wing pads": 7,
    "non wing pads": 7,
    "non-wing pad": 7,
}

# ── SHA-256 HMAC salt (shared with server.py) ──────────────────────
_SALT_FILE = BASE_DIR / ".secret_salt"
if _SALT_FILE.exists():
    _PASSWORD_SALT = _SALT_FILE.read_text(encoding="utf-8").strip()
else:
    _PASSWORD_SALT = secrets.token_hex(16)
    try:
        _SALT_FILE.write_text(_PASSWORD_SALT, encoding="utf-8")
    except Exception:
        pass


def _hash_password(password: str) -> str:
    """Create a salted SHA-256 HMAC hash."""
    return _hmac.new(
        _PASSWORD_SALT.encode("utf-8"),
        password.encode("utf-8"),
        hashlib.sha256,
    ).hexdigest()


def get_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_connection()
    cur = conn.cursor()
    cur.executescript("""
    CREATE TABLE IF NOT EXISTS products (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        details TEXT,
        price REAL NOT NULL,
        slot_number INTEGER NOT NULL UNIQUE,
        capacity INTEGER NOT NULL,
        current_stock INTEGER NOT NULL
    );

    CREATE TABLE IF NOT EXISTS transactions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
        product_id INTEGER,
        quantity INTEGER,
        total_amount REAL,
        payment_method TEXT,
        rfid_user_id INTEGER,
        ir_confirmed INTEGER NOT NULL DEFAULT 1,
        FOREIGN KEY(product_id) REFERENCES products(id)
    );

    CREATE TABLE IF NOT EXISTS rfid_users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        rfid_uid TEXT NOT NULL UNIQUE,
        name TEXT,
        balance REAL NOT NULL DEFAULT 0,
        is_staff INTEGER NOT NULL DEFAULT 0,
        role TEXT NOT NULL DEFAULT 'customer'
    );
    """)
    conn.commit()

    # Ensure new column exists if upgrading from older schema
    cur.execute("PRAGMA table_info(products)")
    product_cols = [row["name"] for row in cur.fetchall()]
    if "details" not in product_cols:
        cur.execute("ALTER TABLE products ADD COLUMN details TEXT")
        conn.commit()

    cur.execute("PRAGMA table_info(transactions)")
    cols = [row["name"] for row in cur.fetchall()]
    if "rfid_user_id" not in cols:
        cur.execute("ALTER TABLE transactions ADD COLUMN rfid_user_id INTEGER")
        conn.commit()

    # Ensure RFID role column exists for door-based access control.
    cur.execute("PRAGMA table_info(rfid_users)")
    user_cols = [row["name"] for row in cur.fetchall()]
    if "role" not in user_cols:
        cur.execute("ALTER TABLE rfid_users ADD COLUMN role TEXT NOT NULL DEFAULT 'customer'")
        conn.commit()

    # Ensure ir_confirmed column exists for tracking vend confirmation status.
    cur.execute("PRAGMA table_info(transactions)")
    txn_cols = [row["name"] for row in cur.fetchall()]
    if "ir_confirmed" not in txn_cols:
        cur.execute("ALTER TABLE transactions ADD COLUMN ir_confirmed INTEGER NOT NULL DEFAULT 1")
        conn.commit()

    # Backfill legacy cards to supported role model.
    # Supported roles: customer, restocker, admin.
    cur.execute("""
        UPDATE rfid_users
        SET role = 'restocker'
        WHERE is_staff = 1 AND (role IS NULL OR role = '' OR role = 'customer' OR role = 'staff')
    """)
    conn.commit()

    # Normalize deprecated maintenance roles to restocker.
    cur.execute(
        """
        UPDATE rfid_users
        SET role = 'restocker'
        WHERE lower(trim(role)) IN ('researcher', 'troubleshooter', 'technician')
        """
    )
    conn.commit()

    # Keep role and is_staff in sync with the supported role model.
    cur.execute(
        """
        UPDATE rfid_users
        SET
            role = CASE
                WHEN lower(trim(role)) = 'admin' THEN 'admin'
                WHEN lower(trim(role)) = 'restocker' THEN 'restocker'
                ELSE 'customer'
            END,
            is_staff = CASE
                WHEN lower(trim(role)) IN ('admin', 'restocker') THEN 1
                ELSE 0
            END
        """
    )
    conn.commit()

    # Seed sample data if empty
    cur.execute("SELECT COUNT(*) AS c FROM products")
    if cur.fetchone()["c"] == 0:
        sample_products = [
            # Slot → Product (per latest physical tray layout)
            ("Alcohol", "Green Cross Isopropyl Alcohol 70% Solution, 60mL", 25.00, 1, 3, 3),
            ("Soap", "Soap, 10grams", 5.00, 2, 7, 7),
            ("Deodorant", "Rexona Shower Clean, 3ml*12packs", 10.00, 3, 9, 9),
            ("Mouthwash", "Scoban Mint Flavor, 10ml*10 packs", 8.00, 4, 7, 7),
            ("Tissues", "Sanicare Hankies, 6 packs", 8.00, 5, 3, 3),
            ("Wet Wipes", "Sanicare Mini Wipes, 6 packs x 8 sheets", 18.00, 6, 3, 3),
            ("Panty Liners", "Charmee Breathable, 20 liners", 5.00, 7, 8, 8),
            ("All Night Pads", "Charmee All Night Plus, 4 pads", 10.00, 8, 6, 6),
            ("Regular W/ Wings Pads", "Charmee Dry Net with wings, 8 pads", 7.00, 9, 7, 7),
            ("Non-Wing Pads", "Charmee Cottony without wings, 8 pads", 7.00, 10, 7, 7),
        ]
        cur.executemany(
            """
            INSERT INTO products (name, details, price, slot_number, capacity, current_stock)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            sample_products,
        )
        conn.commit()
    else:
        # Keep capacities aligned with the latest machine layout even on existing DBs.
        for raw_name, cap in _CAPACITY_RULES.items():
            cur.execute(
                "UPDATE products SET capacity = ? WHERE lower(name) = ?",
                (int(cap), raw_name),
            )
        conn.commit()

        # Defensive: prevent impossible over-cap stock values (e.g., old 10-cap rows).
        cur.execute(
            """
            UPDATE products
            SET current_stock = capacity
            WHERE current_stock > capacity
            """
        )
        conn.commit()

    # Seed a default staff RFID user for testing
    cur.execute("SELECT COUNT(*) AS c FROM rfid_users")
    if cur.fetchone()["c"] == 0:
        cur.execute("""
            INSERT INTO rfid_users (rfid_uid, name, balance, is_staff, role)
            VALUES (?, ?, ?, ?, ?)
        """, ("STAFF001", "Default Staff", 0.0, 1, "restocker"))
        conn.commit()

    # Admin credentials table (single row) with hashed password
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS admin_settings (
            id INTEGER PRIMARY KEY CHECK (id = 1),
            username TEXT NOT NULL,
            password_hash TEXT NOT NULL
        );
        """
    )
    cur.execute("SELECT COUNT(*) AS c FROM admin_settings")
    if cur.fetchone()["c"] == 0:
        default_user = "admin"
        default_pass = "admin"
        pwd_hash = _hash_password(default_pass)
        cur.execute(
            "INSERT INTO admin_settings (id, username, password_hash) VALUES (1, ?, ?)",
            (default_user, pwd_hash),
        )
        conn.commit()

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS hardware_settings (
            setting_key TEXT PRIMARY KEY,
            setting_value TEXT NOT NULL
        );
        """
    )
    cur.execute("SELECT COUNT(*) AS c FROM hardware_settings")
    if cur.fetchone()["c"] == 0:
        cur.executemany(
            "INSERT INTO hardware_settings (setting_key, setting_value) VALUES (?, ?)",
            [
                ("coin_pulse_value", "1.0"),
            ],
        )
        conn.commit()

    conn.close()


def get_admin_credentials():
    """Return (username, password_hash) for the single admin account."""
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT username, password_hash FROM admin_settings WHERE id = 1")
    row = cur.fetchone()
    conn.close()
    if not row:
        return None, None
    return row["username"], row["password_hash"]


def update_admin_credentials(new_username: str, new_password: str):
    """Update the single admin username and password (stored as salted HMAC-SHA256)."""
    pwd_hash = _hash_password(new_password)
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(
        """
        UPDATE admin_settings
        SET username = ?, password_hash = ?
        WHERE id = 1
        """,
        (new_username, pwd_hash),
    )
    conn.commit()
    conn.close()


def get_all_products():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM products ORDER BY slot_number")
    rows = cur.fetchall()
    conn.close()
    return rows


def get_product_by_id(product_id: int):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM products WHERE id = ?", (product_id,))
    row = cur.fetchone()
    conn.close()
    return row


def decrement_stock(product_id: int, quantity: int):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        UPDATE products
        SET current_stock = current_stock - ?
        WHERE id = ? AND current_stock >= ?
    """, (quantity, product_id, quantity))
    if cur.rowcount == 0:
        conn.rollback()
        conn.close()
        raise ValueError("Insufficient stock")
    conn.commit()
    conn.close()


def record_transaction(product_id, quantity, total_amount, payment_method, rfid_user_id=None, ir_confirmed: bool = True):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO transactions (product_id, quantity, total_amount, payment_method, rfid_user_id, ir_confirmed)
        VALUES (?, ?, ?, ?, ?, ?)
    """, (product_id, quantity, total_amount, payment_method, rfid_user_id, 1 if ir_confirmed else 0))
    conn.commit()
    conn.close()


def reset_transactions():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM transactions")
    try:
        cur.execute("DELETE FROM sqlite_sequence WHERE name = 'transactions'")
    except sqlite3.OperationalError:
        pass
    conn.commit()
    conn.close()


def get_user_by_uid(rfid_uid: str):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM rfid_users WHERE rfid_uid = ?", (rfid_uid,))
    row = cur.fetchone()
    conn.close()
    return row


def create_user(
    rfid_uid: str,
    name: str | None = None,
    is_staff: int = 0,
    initial_balance: float = 0.0,
    role: str = "customer",
):
    clean_role = (role or "customer").strip().lower()
    if clean_role not in {"customer", "restocker", "admin"}:
        clean_role = "customer"
    clean_is_staff = 1 if clean_role in {"restocker", "admin"} else 0

    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO rfid_users (rfid_uid, name, balance, is_staff, role)
        VALUES (?, ?, ?, ?, ?)
    """, (rfid_uid, name, initial_balance, clean_is_staff, clean_role))
    conn.commit()
    user_id = cur.lastrowid
    conn.close()
    return user_id


def update_user_balance(user_id: int, new_balance: float):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("UPDATE rfid_users SET balance = ? WHERE id = ?", (new_balance, user_id))
    conn.commit()
    conn.close()


def adjust_user_balance(user_id: int, delta: float):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("UPDATE rfid_users SET balance = balance + ? WHERE id = ?", (delta, user_id))
    conn.commit()
    conn.close()


def get_all_rfid_users():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT id, rfid_uid, name, balance, is_staff, role
        FROM rfid_users
        ORDER BY id
        """
    )
    rows = cur.fetchall()
    conn.close()
    return rows


def update_rfid_user_role(user_id: int, role: str):
    clean_role = (role or "customer").strip().lower()
    if clean_role not in {"customer", "restocker", "admin"}:
        clean_role = "customer"
    is_staff = 1 if clean_role in {"restocker", "admin"} else 0
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(
        "UPDATE rfid_users SET role = ?, is_staff = ? WHERE id = ?",
        (clean_role, is_staff, user_id),
    )
    conn.commit()
    conn.close()


def update_rfid_user(user_id: int, rfid_uid: str, name: str | None, balance: float, role: str):
    clean_role = (role or "customer").strip().lower()
    if clean_role not in {"customer", "restocker", "admin"}:
        clean_role = "customer"
    is_staff = 1 if clean_role in {"restocker", "admin"} else 0
    clean_uid = (rfid_uid or "").strip().upper()
    clean_name = (name or None)
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(
        """
        UPDATE rfid_users
        SET rfid_uid = ?, name = ?, balance = ?, role = ?, is_staff = ?
        WHERE id = ?
        """,
        (clean_uid, clean_name, float(balance), clean_role, is_staff, user_id),
    )
    if cur.rowcount == 0:
        conn.rollback()
        conn.close()
        raise ValueError("RFID user not found")
    conn.commit()
    conn.close()


def delete_rfid_user(user_id: int):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM rfid_users WHERE id = ?", (user_id,))
    if cur.rowcount == 0:
        conn.rollback()
        conn.close()
        raise ValueError("RFID user not found")
    conn.commit()
    conn.close()


def get_hardware_setting(key: str, default: str | None = None):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT setting_value FROM hardware_settings WHERE setting_key = ?", (key,))
    row = cur.fetchone()
    conn.close()
    if not row:
        return default
    return row["setting_value"]


def set_hardware_setting(key: str, value: str):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO hardware_settings (setting_key, setting_value)
        VALUES (?, ?)
        ON CONFLICT(setting_key) DO UPDATE SET setting_value = excluded.setting_value
        """,
        (key, value),
    )
    conn.commit()
    conn.close()


def restock_product(product_id: int, amount: int, new_price: float | None = None):
    conn = get_connection()
    cur = conn.cursor()
    if new_price is not None:
        cur.execute("""
            UPDATE products
            SET current_stock = MIN(capacity, current_stock + ?),
                price = ?
            WHERE id = ?
        """, (amount, new_price, product_id))
    else:
        cur.execute("""
            UPDATE products
            SET current_stock = MIN(capacity, current_stock + ?)
            WHERE id = ?
        """, (amount, product_id))
    conn.commit()
    conn.close()


def export_sales_report():
    """Create an Excel file with all transactions, and daily/monthly summaries."""
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT
            t.timestamp,
            DATE(t.timestamp) AS sale_date,
            strftime('%Y-%m', t.timestamp) AS sale_month,
            p.name AS product_name,
            t.quantity,
            t.total_amount,
            t.payment_method,
            u.rfid_uid
        FROM transactions t
        LEFT JOIN products p ON t.product_id = p.id
        LEFT JOIN rfid_users u ON t.rfid_user_id = u.id
        ORDER BY t.timestamp
    """)
    rows = cur.fetchall()
    conn.close()

    wb = Workbook()
    ws_all = wb.active
    ws_all.title = "All Transactions"
    headers = [
        "Timestamp", "Date", "Month (YYYY-MM)", "Product",
        "Quantity", "Total Amount", "Payment Method", "RFID UID",
    ]
    ws_all.append(headers)

    daily_summary = {}
    monthly_summary = {}

    for r in rows:
        ts = r["timestamp"]
        date = r["sale_date"]
        month = r["sale_month"]
        product = r["product_name"] or ""
        qty = r["quantity"] if r["quantity"] is not None else 0
        total_amt = r["total_amount"] if r["total_amount"] is not None else 0.0
        method = r["payment_method"] or ""
        uid = r["rfid_uid"] or ""

        ws_all.append([ts, date, month, product, qty, total_amt, method, uid])

        if date not in daily_summary:
            daily_summary[date] = {"quantity": 0, "amount": 0.0}
        daily_summary[date]["quantity"] += qty
        daily_summary[date]["amount"] += total_amt

        if month not in monthly_summary:
            monthly_summary[month] = {"quantity": 0, "amount": 0.0}
        monthly_summary[month]["quantity"] += qty
        monthly_summary[month]["amount"] += total_amt

    ws_daily = wb.create_sheet(title="Daily Summary")
    ws_daily.append(["Date", "Total Quantity", "Total Amount"])
    for date, agg in sorted(daily_summary.items()):
        ws_daily.append([date, agg["quantity"], agg["amount"]])

    ws_monthly = wb.create_sheet(title="Monthly Summary")
    ws_monthly.append(["Month (YYYY-MM)", "Total Quantity", "Total Amount"])
    for month, agg in sorted(monthly_summary.items()):
        ws_monthly.append([month, agg["quantity"], agg["amount"]])

    stamp = dt.now().strftime("%Y%m%d_%H%M%S")
    filename = get_reports_dir() / f"sales_report_{stamp}.xlsx"
    wb.save(filename)
    return filename


def get_admin_overview_stats():
    """Compute high-level stats for the admin dashboard."""
    conn = get_connection()
    cur = conn.cursor()

    cur.execute(
        """
        SELECT
            COALESCE(SUM(total_amount), 0) AS total_sales,
            COUNT(*) AS orders
        FROM transactions
        WHERE total_amount IS NOT NULL
        """
    )
    row = cur.fetchone()
    total_sales = row["total_sales"]
    orders = row["orders"]

    cur.execute(
        """
        SELECT COUNT(DISTINCT rfid_user_id) AS active_customers
        FROM transactions
        WHERE rfid_user_id IS NOT NULL
        """
    )
    active_customers = cur.fetchone()["active_customers"]

    # Low-stock thresholds (per final product layout)
    cur.execute("SELECT name, current_stock FROM products")
    rows = cur.fetchall()

    def _threshold(name: str) -> int:
        n = (name or "").strip().lower()
        if n == "alcohol":
            return 1
        if n in ("wet wipes", "wetwipes", "wipes"):
            return 1
        if n in ("tissue", "tissues"):
            return 1
        if n in ("all night pads", "all-night pads"):
            return 2
        return 3

    low_stock = sum(1 for r in rows if int(r["current_stock"] or 0) <= _threshold(r["name"]))

    conn.close()

    return {
        "total_sales": total_sales,
        "orders": orders,
        "active_customers": active_customers,
        "low_stock": low_stock,
    }


def get_ir_confirmation_stats():
    """Return IR confirmation statistics: total transactions and count of unconfirmed."""
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT
            COUNT(*) AS total_txns,
            SUM(CASE WHEN ir_confirmed = 0 THEN 1 ELSE 0 END) AS ir_failed
        FROM transactions
        WHERE payment_method IN ('cash', 'rfid_purchase')
        """
    )
    row = cur.fetchone()
    conn.close()
    total = row["total_txns"] or 0
    failed = row["ir_failed"] or 0
    return {
        "total_dispensed": total,
        "ir_failures": failed,
        "ir_success_rate": (total - failed) / total * 100 if total > 0 else 100.0,
    }


def get_sales_trend_data(days: int = 15):
    """Return daily sales totals for the last N days, including zero-sales days."""
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT DATE(timestamp) AS sale_date, COALESCE(SUM(total_amount), 0) AS total_sales
        FROM transactions
        WHERE DATE(timestamp) >= DATE('now', ?)
        GROUP BY DATE(timestamp)
        ORDER BY DATE(timestamp)
        """,
        (f"-{days - 1} day",),
    )
    rows = cur.fetchall()
    conn.close()

    totals_by_date = {row["sale_date"]: float(row["total_sales"] or 0) for row in rows}
    today = dt.now().date()
    points = []
    for offset in range(days - 1, -1, -1):
        current_date = today - timedelta(days=offset)
        key = current_date.strftime("%Y-%m-%d")
        points.append(
            {
                "date": current_date,
                "label": current_date.strftime("%m/%d"),
                "value": totals_by_date.get(key, 0.0),
            }
        )
    return points


def get_monthly_sales_data(months: int = 6):
    """Return monthly sales totals for the last N months."""
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT strftime('%Y-%m', timestamp) AS sale_month, COALESCE(SUM(total_amount), 0) AS total_sales
        FROM transactions
        WHERE DATE(timestamp) >= DATE('now', 'start of month', ?)
        GROUP BY strftime('%Y-%m', timestamp)
        ORDER BY sale_month
        """,
        (f"-{months - 1} month",),
    )
    rows = cur.fetchall()
    conn.close()

    totals_by_month = {row["sale_month"]: float(row["total_sales"] or 0) for row in rows}
    today = dt.now().date()
    current_month = today.replace(day=1)
    points = []
    for offset in range(months - 1, -1, -1):
        year = current_month.year
        month = current_month.month - offset
        while month <= 0:
            month += 12
            year -= 1
        while month > 12:
            month -= 12
            year += 1
        key = f"{year:04d}-{month:02d}"
        label = f"{month:02d}/{str(year)[2:]}"
        points.append({"label": label, "value": totals_by_month.get(key, 0.0)})
    return points


def get_top_selling_products(limit: int | None = None):
    """Return product sales totals, including products with zero sales."""
    conn = get_connection()
    cur = conn.cursor()
    query = """
        SELECT
            p.name AS product_name,
            COALESCE(SUM(t.quantity), 0) AS total_quantity
        FROM products p
        LEFT JOIN transactions t ON t.product_id = p.id AND t.quantity IS NOT NULL
        GROUP BY p.id, p.name
        ORDER BY total_quantity DESC, p.name ASC
    """
    params: tuple = ()
    if limit is not None:
        query += " LIMIT ?"
        params = (int(limit),)

    cur.execute(query, params)
    rows = cur.fetchall()
    conn.close()
    return [{"label": row["product_name"], "value": float(row["total_quantity"] or 0)} for row in rows]


def get_low_stock_chart_data(limit: int = 10):
    """Return low-stock product data for charting (per-product thresholds)."""
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT name, current_stock, capacity FROM products")
    rows = cur.fetchall()
    conn.close()

    def _threshold(name: str) -> int:
        n = (name or "").strip().lower()
        if n == "alcohol":
            return 1
        if n in ("wet wipes", "wetwipes", "wipes"):
            return 1
        if n in ("tissue", "tissues"):
            return 1
        if n in ("all night pads", "all-night pads"):
            return 2
        return 3

    low = [r for r in rows if int(r["current_stock"] or 0) <= _threshold(r["name"])]
    low.sort(key=lambda r: (int(r["current_stock"] or 0), str(r["name"] or "")))
    low = low[: int(limit)]

    return [
        {"label": row["name"], "value": float(row["current_stock"]), "capacity": float(row["capacity"])}
        for row in low
    ]
