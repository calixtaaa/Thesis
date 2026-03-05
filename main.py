import os
import sys
import time
import uuid
import datetime
import hashlib
import sqlite3
from pathlib import Path
import tkinter as tk
from tkinter import messagebox, simpledialog
from openpyxl import Workbook

from admin.reports import list_sales_reports, open_sales_report

# ======================
#  ENV & GPIO HANDLING
# ======================

ON_RPI = False
try:
    import RPi.GPIO as GPIO  # type: ignore
    ON_RPI = True
except ImportError:
    # Simple mock for development on Windows
    class MockGPIO:
        BCM = "BCM"
        OUT = "OUT"
        IN = "IN"
        HIGH = 1
        LOW = 0
        PUD_UP = "PUD_UP"
        FALLING = "FALLING"

        def setmode(self, *_a, **_k): pass
        def setwarnings(self, *_a, **_k): pass
        def setup(self, *_a, **_k): pass
        def output(self, *_a, **_k): pass
        def cleanup(self): pass
        def add_event_detect(self, *_a, **_k): pass

    GPIO = MockGPIO()  # type: ignore

# ======================
#  CONFIG & CONSTANTS
# ======================

BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / "vending.db"

# Simple theming (light / dark)
THEMES = {
    "light": {
        "bg": "#ffffff",
        "fg": "#000000",
        "button_bg": "#f0f0f0",
        "button_fg": "#000000",
    },
    "dark": {
        "bg": "#202124",
        "fg": "#e8eaed",
        "button_bg": "#3c4043",
        "button_fg": "#e8eaed",
    },
}

# Simple pin assignments (adjust on real Pi)
PRODUCT_STEPPER_PINS = {
    1: {"step": 17, "dir": 27},   # Slot 1
    2: {"step": 22, "dir": 23},   # Slot 2
}
COIN_HOPPER_PIN = 24
STEPS_PER_PRODUCT = 200  # adjust per mechanism
COINS_PER_SECOND = 5     # for hopper simulation
COIN_VALUE = 1.0         # 1 peso per coin (example)


# ======================
#  DATABASE LAYER
# ======================

def get_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_connection()
    cur = conn.cursor()
    cur.executescript("""
    CREATE TABLE IF NOT EXISTS products (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        price REAL NOT NULL,
        slot_number INTEGER NOT NULL UNIQUE,
        capacity INTEGER NOT NULL,
        current_stock INTEGER NOT NULL
    );

    CREATE TABLE IF NOT EXISTS transactions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
        product_id INTEGER,
        quantity INTEGER,
        total_amount REAL,
        payment_method TEXT,
        rfid_user_id INTEGER,
        FOREIGN KEY(product_id) REFERENCES products(id)
    );

    CREATE TABLE IF NOT EXISTS rfid_users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        rfid_uid TEXT NOT NULL UNIQUE,
        name TEXT,
        balance REAL NOT NULL DEFAULT 0,
        is_staff INTEGER NOT NULL DEFAULT 0
    );
    """)
    conn.commit()

    # Ensure new column exists if upgrading from older schema
    cur.execute("PRAGMA table_info(transactions)")
    cols = [row["name"] for row in cur.fetchall()]
    if "rfid_user_id" not in cols:
        cur.execute("ALTER TABLE transactions ADD COLUMN rfid_user_id INTEGER")
        conn.commit()

    # Seed sample data if empty
    cur.execute("SELECT COUNT(*) AS c FROM products")
    if cur.fetchone()["c"] == 0:
        # Approximate SRP prices in PHP for each product in the Philippines
        sample_products = [
            ("All Night Pads",          12.0, 1, 10, 5),
            ("Panty Liners",            8.0,  2, 10, 5),
            ("Regular W/ Wings Pads",   10.0, 3, 10, 5),
            ("Non-Wing Pads",           9.0,  4, 10, 5),
            ("Alcohol",                 25.0, 5, 10, 5),
            ("Mouthwash",               35.0, 6, 10, 5),
            ("Tissues",                 15.0, 7, 10, 5),
            ("Wet Wipes",               25.0, 8, 10, 5),
            ("Deodorant",               50.0, 9, 10, 5),
            ("Soap",                    30.0, 10, 10, 5),
        ]
        cur.executemany(
            """
            INSERT INTO products (name, price, slot_number, capacity, current_stock)
            VALUES (?, ?, ?, ?, ?)
            """,
            sample_products,
        )
        conn.commit()

    # Seed a default staff RFID user for testing
    cur.execute("SELECT COUNT(*) AS c FROM rfid_users")
    if cur.fetchone()["c"] == 0:
        cur.execute("""
            INSERT INTO rfid_users (rfid_uid, name, balance, is_staff)
            VALUES (?, ?, ?, ?)
        """, ("STAFF001", "Default Staff", 0.0, 1))
        conn.commit()

    # Admin credentials table (single row) with hashed password
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS admin_settings (
            id INTEGER PRIMARY KEY CHECK (id = 1),
            username TEXT NOT NULL,
            password_hash TEXT NOT NULL
        );
        """
    )
    cur.execute("SELECT COUNT(*) AS c FROM admin_settings")
    if cur.fetchone()["c"] == 0:
        # Default admin/admin (hashed)
        default_user = "admin"
        default_pass = "admin"
        pwd_hash = hashlib.sha256(default_pass.encode("utf-8")).hexdigest()
        cur.execute(
            "INSERT INTO admin_settings (id, username, password_hash) VALUES (1, ?, ?)",
            (default_user, pwd_hash),
        )
        conn.commit()

    conn.close()


def get_admin_credentials():
    """Return (username, password_hash) for the single admin account."""
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT username, password_hash FROM admin_settings WHERE id = 1")
    row = cur.fetchone()
    conn.close()
    if not row:
        return None, None
    return row["username"], row["password_hash"]


def update_admin_credentials(new_username: str, new_password: str):
    """Update the single admin username and password (stored as SHA-256 hash)."""
    pwd_hash = hashlib.sha256(new_password.encode("utf-8")).hexdigest()
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(
        """
        UPDATE admin_settings
        SET username = ?, password_hash = ?
        WHERE id = 1
        """,
        (new_username, pwd_hash),
    )
    conn.commit()
    conn.close()

def get_all_products():
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM products ORDER BY slot_number")
    rows = cur.fetchall()
    conn.close()
    return rows


def get_product_by_id(product_id: int):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM products WHERE id = ?", (product_id,))
    row = cur.fetchone()
    conn.close()
    return row

def decrement_stock(product_id: int, quantity: int):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        UPDATE products
        SET current_stock = current_stock - ?
        WHERE id = ? AND current_stock >= ?
    """, (quantity, product_id, quantity))
    if cur.rowcount == 0:
        conn.rollback()
        conn.close()
        raise ValueError("Insufficient stock")
    conn.commit()
    conn.close()

def record_transaction(product_id, quantity, total_amount, payment_method, rfid_user_id=None):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO transactions (product_id, quantity, total_amount, payment_method, rfid_user_id)
        VALUES (?, ?, ?, ?, ?)
    """, (product_id, quantity, total_amount, payment_method, rfid_user_id))
    conn.commit()
    conn.close()


def get_user_by_uid(rfid_uid: str):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM rfid_users WHERE rfid_uid = ?", (rfid_uid,))
    row = cur.fetchone()
    conn.close()
    return row


def create_user(rfid_uid: str, name: str | None = None, is_staff: int = 0, initial_balance: float = 0.0):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO rfid_users (rfid_uid, name, balance, is_staff)
        VALUES (?, ?, ?, ?)
    """, (rfid_uid, name, initial_balance, is_staff))
    conn.commit()
    user_id = cur.lastrowid
    conn.close()
    return user_id


def update_user_balance(user_id: int, new_balance: float):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("UPDATE rfid_users SET balance = ? WHERE id = ?", (new_balance, user_id))
    conn.commit()
    conn.close()


def adjust_user_balance(user_id: int, delta: float):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("UPDATE rfid_users SET balance = balance + ? WHERE id = ?", (delta, user_id))
    conn.commit()
    conn.close()


def restock_product(product_id: int, amount: int, new_price: float | None = None):
    conn = get_connection()
    cur = conn.cursor()
    if new_price is not None:
        cur.execute("""
            UPDATE products
            SET current_stock = MIN(capacity, current_stock + ?),
                price = ?
            WHERE id = ?
        """, (amount, new_price, product_id))
    else:
        cur.execute("""
            UPDATE products
            SET current_stock = MIN(capacity, current_stock + ?)
            WHERE id = ?
        """, (amount, product_id))
    conn.commit()
    conn.close()


def export_sales_report():
    """Create an Excel file with all transactions, and daily/monthly summaries."""
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT
            t.timestamp,
            DATE(t.timestamp) AS sale_date,
            strftime('%Y-%m', t.timestamp) AS sale_month,
            p.name AS product_name,
            t.quantity,
            t.total_amount,
            t.payment_method,
            u.rfid_uid
        FROM transactions t
        LEFT JOIN products p ON t.product_id = p.id
        LEFT JOIN rfid_users u ON t.rfid_user_id = u.id
        ORDER BY t.timestamp
    """)
    rows = cur.fetchall()
    conn.close()

    # Prepare Excel workbook
    wb = Workbook()

    # Sheet 1: All Transactions
    ws_all = wb.active
    ws_all.title = "All Transactions"
    headers = [
        "Timestamp", "Date", "Month (YYYY-MM)", "Product",
        "Quantity", "Total Amount", "Payment Method", "RFID UID",
    ]
    ws_all.append(headers)

    daily_summary = {}
    monthly_summary = {}

    for r in rows:
        ts = r["timestamp"]
        date = r["sale_date"]
        month = r["sale_month"]
        product = r["product_name"] or ""
        qty = r["quantity"] if r["quantity"] is not None else 0
        total_amt = r["total_amount"] if r["total_amount"] is not None else 0.0
        method = r["payment_method"] or ""
        uid = r["rfid_uid"] or ""

        ws_all.append([ts, date, month, product, qty, total_amt, method, uid])

        # Build daily summary
        if date not in daily_summary:
            daily_summary[date] = {"quantity": 0, "amount": 0.0}
        daily_summary[date]["quantity"] += qty
        daily_summary[date]["amount"] += total_amt

        # Build monthly summary
        if month not in monthly_summary:
            monthly_summary[month] = {"quantity": 0, "amount": 0.0}
        monthly_summary[month]["quantity"] += qty
        monthly_summary[month]["amount"] += total_amt

    # Sheet 2: Daily Summary
    ws_daily = wb.create_sheet(title="Daily Summary")
    ws_daily.append(["Date", "Total Quantity", "Total Amount"])
    for date, agg in sorted(daily_summary.items()):
        ws_daily.append([date, agg["quantity"], agg["amount"]])

    # Sheet 3: Monthly Summary
    ws_monthly = wb.create_sheet(title="Monthly Summary")
    ws_monthly.append(["Month (YYYY-MM)", "Total Quantity", "Total Amount"])
    for month, agg in sorted(monthly_summary.items()):
        ws_monthly.append([month, agg["quantity"], agg["amount"]])

    # Save file
    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = BASE_DIR / f"sales_report_{stamp}.xlsx"
    wb.save(filename)
    return filename


def get_admin_overview_stats():
    """Compute high-level stats for the admin dashboard."""
    conn = get_connection()
    cur = conn.cursor()

    # Total sales amount and number of orders (completed transactions with amount)
    cur.execute(
        """
        SELECT
            COALESCE(SUM(total_amount), 0) AS total_sales,
            COUNT(*) AS orders
        FROM transactions
        WHERE total_amount IS NOT NULL
        """
    )
    row = cur.fetchone()
    total_sales = row["total_sales"]
    orders = row["orders"]

    # Active customers: distinct RFID users that have at least one transaction
    cur.execute(
        """
        SELECT COUNT(DISTINCT rfid_user_id) AS active_customers
        FROM transactions
        WHERE rfid_user_id IS NOT NULL
        """
    )
    active_customers = cur.fetchone()["active_customers"]

    # Low-stock products (stock <= 20% of capacity)
    cur.execute("SELECT COUNT(*) AS low_count FROM products WHERE capacity > 0 AND current_stock <= capacity * 0.2")
    low_stock = cur.fetchone()["low_count"]

    conn.close()

    return {
        "total_sales": total_sales,
        "orders": orders,
        "active_customers": active_customers,
        "low_stock": low_stock,
    }


# ======================
#  HARDWARE ABSTRACTION
# ======================

def gpio_init():
    GPIO.setwarnings(False)
    GPIO.setmode(GPIO.BCM)

    # Stepper pins
    for cfg in PRODUCT_STEPPER_PINS.values():
        GPIO.setup(cfg["step"], GPIO.OUT)
        GPIO.setup(cfg["dir"], GPIO.OUT)

    # Hopper
    GPIO.setup(COIN_HOPPER_PIN, GPIO.OUT)

def dispense_from_slot(slot_number: int, quantity: int = 1):
    pins = PRODUCT_STEPPER_PINS.get(slot_number)
    if not pins:
        print(f"[HW] No stepper pins configured for slot {slot_number}")
        return

    print(f"[HW] Dispensing from slot {slot_number} x{quantity}")
    steps = STEPS_PER_PRODUCT * quantity
    delay = 0.001  # 1ms

    GPIO.output(pins["dir"], GPIO.HIGH)
    for _ in range(steps):
        GPIO.output(pins["step"], GPIO.HIGH)
        time.sleep(delay)
        GPIO.output(pins["step"], GPIO.LOW)
        time.sleep(delay)

def dispense_change(amount: float):
    coins_to_dispense = int(amount // COIN_VALUE)
    if coins_to_dispense <= 0:
        return
    print(f"[HW] Dispensing change: {coins_to_dispense} coins (₱{amount:.2f})")
    duration = coins_to_dispense / COINS_PER_SECOND
    GPIO.output(COIN_HOPPER_PIN, GPIO.HIGH)
    time.sleep(duration)
    GPIO.output(COIN_HOPPER_PIN, GPIO.LOW)


# ======================
#  SIMPLE CASH SESSION
#  (Simulated for now)
# ======================

class CashSession:
    """
    For now: simulated cash input using "+" buttons on GUI.
    On real machine, replace with pulse‑based reader.
    """
    def __init__(self):
        self.total_amount = 0.0

    def reset(self):
        self.total_amount = 0.0

    def add(self, value: float):
        self.total_amount += value

    def get_amount(self) -> float:
        return self.total_amount

cash_session = CashSession()


# ======================
#  TKINTER GUI
# ======================

class MainApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Hygiene Vending Machine")

        # On the Raspberry Pi 7\" touchscreen, use fullscreen.
        # On desktop (development), use a resizable 800x480 window.
        if ON_RPI:
            self.attributes("-fullscreen", True)
        else:
            self.geometry("800x480")
            self.minsize(800, 480)

        self.current_product = None
        self.current_quantity = 1

        # Theme state
        self.current_theme_name = "light"
        self.current_theme = THEMES[self.current_theme_name]
        self.configure(bg=self.current_theme["bg"])

        # Icon images (loaded lazily if files exist)
        self.staff_icon = None
        self.admin_icon = None
        self.menu_icon = None
        self.search_icon = None

        # Staff icon
        try:
            for name in ("staff.png", "staff_icon.png"):
                staff_path = BASE_DIR / "images" / name
                if staff_path.exists():
                    self.staff_icon = tk.PhotoImage(file=str(staff_path))
                    break
        except Exception:
            self.staff_icon = None

        # Admin icon
        try:
            for name in ("admin.png", "admin_icon.png"):
                admin_path = BASE_DIR / "images" / name
                if admin_path.exists():
                    self.admin_icon = tk.PhotoImage(file=str(admin_path))
                    break
        except Exception:
            self.admin_icon = None

        # Hamburger / menu icon
        try:
            menu_path = BASE_DIR / "images" / "hamburger.png"
            if menu_path.exists():
                # Load and aggressively downscale so it appears as a small navbar icon
                img = tk.PhotoImage(file=str(menu_path))
                # Adjust subsample factors if you later change the SVG export size
                self.menu_icon = img.subsample(4, 4)
        except Exception:
            self.menu_icon = None

        # Search / magnifying glass icon
        try:
            search_path = BASE_DIR / "images" / "magnifying glass.png"
            if search_path.exists():
                img = tk.PhotoImage(file=str(search_path))
                self.search_icon = img.subsample(4, 4)
        except Exception:
            self.search_icon = None

        # Backspace icon for clearing search text
        self.backspace_icon = None
        try:
            backspace_path = BASE_DIR / "images" / "backspace.png"
            if backspace_path.exists():
                img = tk.PhotoImage(file=str(backspace_path))
                self.backspace_icon = img.subsample(4, 4)
        except Exception:
            self.backspace_icon = None

        # Sidebar / navbar state
        self.role_menu = None
        self.sidebar_frame = None

        self.search_var = tk.StringVar()

        self.build_main_menu()

    # ---------- Screen helpers ----------

    def clear_screen(self):
        for w in self.winfo_children():
            w.destroy()

    def apply_theme_to_widget(self, widget):
        """Apply current theme colors to a widget and its children."""
        try:
            widget.configure(bg=self.current_theme["bg"])
        except tk.TclError:
            pass
        for child in widget.winfo_children():
            if isinstance(child, (tk.Button, tk.Label, tk.Frame)):
                try:
                    if isinstance(child, tk.Button):
                        child.configure(
                            bg=self.current_theme["button_bg"],
                            fg=self.current_theme["button_fg"],
                            activebackground=self.current_theme["button_bg"],
                            activeforeground=self.current_theme["button_fg"],
                        )
                    elif isinstance(child, tk.Label) or isinstance(child, tk.Frame):
                        child.configure(
                            bg=self.current_theme["bg"],
                            fg=self.current_theme["fg"] if isinstance(child, tk.Label) else None,
                        )
                except tk.TclError:
                    pass

    def toggle_theme(self):
        """Switch between light and dark modes and refresh current screen."""
        self.current_theme_name = "dark" if self.current_theme_name == "light" else "light"
        self.current_theme = THEMES[self.current_theme_name]
        self.configure(bg=self.current_theme["bg"])
        self.apply_theme_to_widget(self)

    def add_theme_toggle_footer(self):
        """Add a bottom bar with a theme toggle button to the current screen."""
        bottom = tk.Frame(self, bg=self.current_theme["bg"])
        bottom.pack(side=tk.BOTTOM, fill=tk.X, pady=5)
        # Trademark / footer label
        tk.Label(
            bottom,
            text="SyntaxError™",
            font=("Arial", 9),
            bg=self.current_theme["bg"],
            fg=self.current_theme["fg"],
        ).pack(side=tk.LEFT, padx=10)
        tk.Button(
            bottom,
            text=f"Theme: {self.current_theme_name.capitalize()}",
            command=self.toggle_theme,
            bg=self.current_theme["button_bg"],
            fg=self.current_theme["button_fg"],
        ).pack(side=tk.RIGHT, padx=10)
        self.apply_theme_to_widget(self)

    def show_help_dialog(self):
        """Explain the basic usage steps to the user."""
        message = (
            "How to use the vending machine:\n\n"
            "1. Select Product – Tap the item you want.\n"
            "2. Quantity – Choose how many pieces you need.\n"
            "3. Payment – Pay with Cash or RFID Card, then wait for the product to be dispensed.\n\n"
            "You can also buy a new RFID card or reload an existing card using the buttons below."
        )
        messagebox.showinfo("How to use", message)

    def show_role_menu(self):
        """Toggle a left sidebar for Staff/Admin actions (hamburger menu)."""
        # If sidebar already open, close it
        if self.sidebar_frame is not None and self.sidebar_frame.winfo_exists():
            self.sidebar_frame.destroy()
            self.sidebar_frame = None
            return

        # Create a vertical sidebar on the left, similar to the reference image
        sidebar_width = 200
        sidebar = tk.Frame(self, bg=self.current_theme["bg"], width=sidebar_width)
        sidebar.place(x=0, y=0, relheight=1.0)

        tk.Label(
            sidebar,
            text="Menu",
            font=("Arial", 14, "bold"),
            bg=self.current_theme["bg"],
            fg=self.current_theme["fg"],
        ).pack(anchor="w", padx=12, pady=(12, 8))

        def make_nav_button(text, command):
            btn = tk.Button(
                sidebar,
                text=text,
                anchor="w",
                font=("Arial", 11),
                command=command,
                bg=self.current_theme["button_bg"],
                fg=self.current_theme["button_fg"],
                relief="flat",
                padx=10,
                pady=4,
            )
            btn.pack(fill=tk.X, padx=8, pady=2)
            return btn

        # Dashboard just closes sidebar and keeps current screen
        make_nav_button("Dashboard (user view)", lambda: self.show_role_menu())

        # Staff restock screen
        make_nav_button(
            "Staff – Restock products",
            lambda: (self.show_role_menu(), self.enter_restock_mode()),
        )

        # Admin dashboard
        make_nav_button(
            "Admin – Reports & settings",
            lambda: (self.show_role_menu(), self.enter_admin_dashboard()),
        )

        # Sign out/back to main
        make_nav_button(
            "Back to main screen",
            lambda: (self.show_role_menu(), self.build_main_menu()),
        )

        self.apply_theme_to_widget(sidebar)
        self.sidebar_frame = sidebar

    # ---------- Main Menu ----------

    def build_main_menu(self):
        self.clear_screen()
        all_products = get_all_products()

        top = tk.Frame(self, bg=self.current_theme["bg"])
        top.pack(side=tk.TOP, fill=tk.X)

        header = tk.Frame(top, bg=self.current_theme["bg"])
        header.pack(side=tk.TOP, fill=tk.X, pady=2, padx=8)

        title_block = tk.Frame(header, bg=self.current_theme["bg"])
        title_block.pack(side=tk.LEFT)
        tk.Label(
            title_block,
            text="Select Product",
            font=("Arial", 20, "bold"),
            bg=self.current_theme["bg"],
            fg=self.current_theme["fg"],
        ).pack(anchor="w")
        tk.Label(
            title_block,
            text="Step 1 of 3 – Tap a product to begin",
            font=("Arial", 10),
            bg=self.current_theme["bg"],
            fg=self.current_theme["fg"],
        ).pack(anchor="w")

        # Hamburger menu icon on the top-right
        icons_frame = tk.Frame(header, bg=self.current_theme["bg"])
        icons_frame.pack(side=tk.RIGHT)
        if self.menu_icon is not None:
            tk.Button(
                icons_frame,
                image=self.menu_icon,
                command=self.show_role_menu,
                bg=self.current_theme["bg"],
                activebackground=self.current_theme["bg"],
                bd=0,
                highlightthickness=0,
                width=self.menu_icon.width(),
                height=self.menu_icon.height(),
                padx=0,
                pady=0,
            ).pack(side=tk.RIGHT, padx=4, pady=0)
        else:
            tk.Button(
                icons_frame,
                text="☰",
                command=self.show_role_menu,
                bg=self.current_theme["button_bg"],
                fg=self.current_theme["button_fg"],
            ).pack(side=tk.RIGHT, padx=2, pady=0)

        # Search bar row (pill-shaped style)
        search_row = tk.Frame(self, bg=self.current_theme["bg"])
        search_row.pack(side=tk.TOP, fill=tk.X, padx=12, pady=(0, 8))

        # Outer container to simulate rounded, light background
        search_bg = "#f1f3f4"
        pill = tk.Frame(
            search_row,
            bg=search_bg,
            bd=0,
            highlightthickness=0,
        )
        pill.pack(side=tk.LEFT, fill=tk.X, expand=True)

        if self.search_icon is not None:
            tk.Label(
                pill,
                image=self.search_icon,
                bg=search_bg,
            ).pack(side=tk.LEFT, padx=(8, 4), pady=4)

        search_entry = tk.Entry(
            pill,
            textvariable=self.search_var,
            font=("Arial", 11),
            bd=0,
            highlightthickness=0,
            relief="flat",
            bg=search_bg,
        )
        search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 4), pady=4)
        if not self.search_var.get():
            self.search_var.set("Search for products")

        def on_search_focus_in(_event):
            if self.search_var.get() == "Search for products":
                self.search_var.set("")

        def on_search_focus_out(_event):
            if not self.search_var.get().strip():
                self.search_var.set("Search for products")

        def refresh_products(_event=None):
            self.build_main_menu()

        def clear_one_char():
            current = self.search_var.get()
            if current == "Search for products" or not current:
                return
            self.search_var.set(current[:-1])
            refresh_products()

        search_entry.bind("<FocusIn>", on_search_focus_in)
        search_entry.bind("<FocusOut>", on_search_focus_out)
        search_entry.bind("<KeyRelease>", refresh_products)

        # Backspace icon button to delete one character from search
        if self.backspace_icon is not None:
            tk.Button(
                pill,
                image=self.backspace_icon,
                command=clear_one_char,
                bg=search_bg,
                activebackground=search_bg,
                bd=0,
                highlightthickness=0,
            ).pack(side=tk.LEFT, padx=(0, 8), pady=4)

        # Filter products based on search text
        query = self.search_var.get().strip()
        if query and query != "Search for products":
            products = [p for p in all_products if query.lower() in p["name"].lower()]
        else:
            products = all_products

        # Scrollable product area
        content_frame = tk.Frame(self, bg=self.current_theme["bg"])
        content_frame.pack(expand=True, fill=tk.BOTH)

        canvas = tk.Canvas(
            content_frame,
            bg=self.current_theme["bg"],
            highlightthickness=0,
        )
        scrollbar = tk.Scrollbar(content_frame, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)

        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        grid = tk.Frame(canvas, bg=self.current_theme["bg"])
        canvas.create_window((0, 0), window=grid, anchor="nw")

        # Make the product grid responsive (3 columns)
        for c in range(3):
            grid.grid_columnconfigure(c, weight=1)

        for idx, p in enumerate(products):
            name = p["name"]
            stock = p["current_stock"]
            capacity = p["capacity"]
            price = p["price"]

            btn_text = f"{name}\n{stock}/{capacity}\n₱{price:.2f}"
            state = tk.NORMAL if stock > 0 else tk.DISABLED

            btn = tk.Button(
                grid,
                text=btn_text,
                state=state,
                command=lambda prod=p: self.select_product(prod),
                bg=self.current_theme["button_bg"],
                fg=self.current_theme["button_fg"],
                activebackground=self.current_theme["button_bg"],
                activeforeground=self.current_theme["button_fg"],
                wraplength=200,
                padx=10,
                pady=10,
            )
            r = idx // 3
            c = idx % 3
            btn.grid(row=r, column=c, padx=10, pady=10, sticky="nsew")
            grid.grid_rowconfigure(r, weight=1)

        def _on_frame_configure(_event):
            canvas.configure(scrollregion=canvas.bbox("all"))

        grid.bind("<Configure>", _on_frame_configure)

        # Mouse wheel scrolling for better UX
        def _on_mousewheel(event):
            # Windows / Mac delta handling
            delta = -1 * int(event.delta / 120)
            canvas.yview_scroll(delta, "units")

        canvas.bind_all("<MouseWheel>", _on_mousewheel)

        bottom = tk.Frame(self, bg=self.current_theme["bg"])
        bottom.pack(side=tk.BOTTOM, fill=tk.X, pady=10)

        # Visible buttons for regular users
        tk.Button(
            bottom,
            text="Reload (RFID)",
            command=self.reload_card_flow,
            bg="#4CAF50",
            fg=self.current_theme["button_fg"],
        ).pack(side=tk.LEFT, padx=10)
        tk.Button(
            bottom,
            text="Buy RFID Card",
            command=self.buy_card_flow,
            bg="#4CAF50",
            fg=self.current_theme["button_fg"],
        ).pack(side=tk.LEFT, padx=10)

        # Simple help button explaining the 3 steps
        tk.Button(
            bottom,
            text="How to use?",
            command=self.show_help_dialog,
            bg=self.current_theme["button_bg"],
            fg=self.current_theme["button_fg"],
        ).pack(side=tk.LEFT, padx=10)

        # Trademark footer label
        tk.Label(
            bottom,
            text="SyntaxError™",
            font=("Arial", 9),
            bg=self.current_theme["bg"],
            fg=self.current_theme["fg"],
        ).pack(side=tk.LEFT, padx=10)

        tk.Button(
            bottom,
            text=f"Theme: {self.current_theme_name.capitalize()}",
            command=self.toggle_theme,
            bg=self.current_theme["button_bg"],
            fg=self.current_theme["button_fg"],
        ).pack(side=tk.RIGHT, padx=10)

        self.apply_theme_to_widget(self)

    def select_product(self, product_row):
        self.current_product = product_row
        self.current_quantity = 1
        self.show_quantity_screen()

    # ---------- Quantity Screen ----------

    def show_quantity_screen(self):
        self.clear_screen()
        p = self.current_product

        frame = tk.Frame(self, bg=self.current_theme["bg"])
        frame.pack(expand=True, fill=tk.BOTH)

        tk.Label(
            frame,
            text="Step 2 of 3 – Choose quantity",
            font=("Arial", 12),
            bg=self.current_theme["bg"],
            fg=self.current_theme["fg"],
        ).pack(pady=(10, 0))

        tk.Label(frame, text=f"Selected: {p['name']}", font=("Arial", 18)).pack(pady=10)
        tk.Label(frame, text=f"Price: ₱{p['price']:.2f}", font=("Arial", 14)).pack(pady=5)
        tk.Label(frame, text=f"Available: {p['current_stock']}", font=("Arial", 12)).pack(pady=5)

        qty_var = tk.IntVar(value=self.current_quantity)

        def update_qty(delta):
            new = qty_var.get() + delta
            if 1 <= new <= p["current_stock"]:
                qty_var.set(new)

        qty_frame = tk.Frame(frame, bg=self.current_theme["bg"])
        qty_frame.pack(pady=20)

        tk.Button(qty_frame, text="-", width=5, command=lambda: update_qty(-1)).pack(side=tk.LEFT, padx=10)
        tk.Label(qty_frame, textvariable=qty_var, font=("Arial", 24)).pack(side=tk.LEFT, padx=10)
        tk.Button(qty_frame, text="+", width=5, command=lambda: update_qty(1)).pack(side=tk.LEFT, padx=10)

        def proceed():
            self.current_quantity = qty_var.get()
            self.show_payment_method_screen()

        tk.Button(frame, text="Continue to payment", font=("Arial", 14), command=proceed).pack(pady=10)
        tk.Button(frame, text="Back to products", font=("Arial", 12), command=self.build_main_menu).pack(pady=5)

        self.add_theme_toggle_footer()

    # ---------- Payment Method ----------

    def show_payment_method_screen(self):
        self.clear_screen()
        p = self.current_product
        q = self.current_quantity
        total = p["price"] * q

        frame = tk.Frame(self, bg=self.current_theme["bg"])
        frame.pack(expand=True, fill=tk.BOTH)

        tk.Label(
            frame,
            text="Step 3 of 3 – Choose payment method",
            font=("Arial", 12),
            bg=self.current_theme["bg"],
            fg=self.current_theme["fg"],
        ).pack(pady=(10, 0))
        tk.Label(frame, text="Choose Payment Method", font=("Arial", 20)).pack(pady=5)
        tk.Label(frame, text=f"Product: {p['name']} x{q}", font=("Arial", 14)).pack(pady=5)
        tk.Label(frame, text=f"Total: ₱{total:.2f}", font=("Arial", 18)).pack(pady=10)

        tk.Button(
            frame,
            text="Pay with Cash (insert coins/bills)",
            width=20,
            height=2,
            command=lambda: self.cash_payment_flow(total)
        ).pack(pady=10)

        tk.Button(
            frame,
            text="Pay with RFID Card (cashless)",
            width=25,
            height=2,
            command=lambda: self.rfid_payment_flow(total)
        ).pack(pady=10)

        tk.Button(frame, text="Back to quantity", command=self.show_quantity_screen).pack(pady=10)

        self.add_theme_toggle_footer()

    # ---------- Cash Payment Flow ----------

    def cash_payment_flow(self, total_amount: float):
        self.clear_screen()
        cash_session.reset()

        frame = tk.Frame(self, bg=self.current_theme["bg"])
        frame.pack(expand=True, fill=tk.BOTH)

        tk.Label(
            frame,
            text="Pay with Cash",
            font=("Arial", 20),
        ).pack(pady=(10, 0))
        tk.Label(
            frame,
            text="Press the buttons below to simulate inserting coins/bills.",
            font=("Arial", 10),
        ).pack(pady=2)
        tk.Label(frame, text=f"Total to Pay: ₱{total_amount:.2f}", font=("Arial", 14)).pack(pady=5)

        amount_var = tk.DoubleVar(value=0.0)
        remaining_var = tk.DoubleVar(value=total_amount)

        tk.Label(frame, text="Amount Inserted:", font=("Arial", 14)).pack(pady=5)
        tk.Label(frame, textvariable=amount_var, font=("Arial", 22)).pack()

        tk.Label(frame, text="Remaining:", font=("Arial", 14)).pack(pady=5)
        tk.Label(frame, textvariable=remaining_var, font=("Arial", 22)).pack()

        # Simulated cash buttons for development
        btn_frame = tk.Frame(frame, bg=self.current_theme["bg"])
        btn_frame.pack(pady=15)

        def add_money(val):
            cash_session.add(val)
            current = cash_session.get_amount()
            amount_var.set(current)
            remaining_var.set(max(0.0, total_amount - current))

        tk.Button(btn_frame, text="+₱1", width=8, command=lambda: add_money(1)).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="+₱5", width=8, command=lambda: add_money(5)).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="+₱10", width=8, command=lambda: add_money(10)).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="+₱20", width=8, command=lambda: add_money(20)).pack(side=tk.LEFT, padx=5)

        def finish_if_enough():
            current = cash_session.get_amount()
            if current < total_amount:
                messagebox.showwarning("Not enough", "Please insert more cash.")
                return
            self.complete_purchase_cash(total_amount, current)

        tk.Button(
            frame,
            text="Dispense product",
            font=("Arial", 14),
            command=finish_if_enough,
            bg="#4CAF50",
            fg=self.current_theme["button_fg"],
        ).pack(pady=15)

        tk.Button(
            frame,
            text="Cancel and go back",
            command=self.build_main_menu,
            bg="#E53935",
            fg=self.current_theme["button_fg"],
        ).pack(pady=5)

        self.add_theme_toggle_footer()

    def complete_purchase_cash(self, total_amount: float, inserted: float):
        p = self.current_product
        q = self.current_quantity
        change = max(0.0, inserted - total_amount)

        self.show_wait_screen("Processing payment and dispensing...")

        try:
            decrement_stock(p["id"], q)
            dispense_from_slot(p["slot_number"], q)
            if change > 0:
                dispense_change(change)
            record_transaction(p["id"], q, total_amount, "cash")
            messagebox.showinfo(
                "Thank you",
                f"Please take your product.\nChange: ₱{change:.2f}"
            )
        except Exception as e:
            messagebox.showerror("Error", str(e))
        finally:
            self.build_main_menu()

    # ---------- Utility Screens ----------

    def show_wait_screen(self, text: str):
        self.clear_screen()
        frame = tk.Frame(self, bg=self.current_theme["bg"])
        frame.pack(expand=True, fill=tk.BOTH)
        tk.Label(frame, text=text, font=("Arial", 18)).pack(pady=20)

        self.add_theme_toggle_footer()

    # ---------- RFID Card Purchase ----------

    def buy_card_flow(self):
        """Customer buys a new RFID card for a fixed price (e.g. ₱50)."""
        CARD_PRICE = 50.0
        self.clear_screen()
        cash_session.reset()

        frame = tk.Frame(self, bg=self.current_theme["bg"])
        frame.pack(expand=True, fill=tk.BOTH)

        tk.Label(frame, text="Buy a New RFID Card", font=("Arial", 20)).pack(pady=10)
        tk.Label(frame, text=f"Please pay ₱{CARD_PRICE:.2f}", font=("Arial", 14)).pack(pady=5)

        amount_var = tk.DoubleVar(value=0.0)
        remaining_var = tk.DoubleVar(value=CARD_PRICE)

        tk.Label(frame, text="Amount Inserted:", font=("Arial", 14)).pack(pady=5)
        tk.Label(frame, textvariable=amount_var, font=("Arial", 22)).pack()

        tk.Label(frame, text="Remaining:", font=("Arial", 14)).pack(pady=5)
        tk.Label(frame, textvariable=remaining_var, font=("Arial", 22)).pack()

        btn_frame = tk.Frame(frame, bg=self.current_theme["bg"])
        btn_frame.pack(pady=15)

        def add_money(val):
            cash_session.add(val)
            current = cash_session.get_amount()
            amount_var.set(current)
            remaining_var.set(max(0.0, CARD_PRICE - current))

        tk.Button(btn_frame, text="+₱1", width=8, command=lambda: add_money(1)).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="+₱5", width=8, command=lambda: add_money(5)).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="+₱10", width=8, command=lambda: add_money(10)).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="+₱20", width=8, command=lambda: add_money(20)).pack(side=tk.LEFT, padx=5)

        def confirm_purchase():
            inserted = cash_session.get_amount()
            if inserted < CARD_PRICE:
                messagebox.showwarning("Not enough", "Please insert full card price.")
                return

            self.show_wait_screen("Issuing RFID card...")

            # Generate a simple random UID for demonstration
            uid = uuid.uuid4().hex[:8].upper()
            user_id = create_user(uid, name=None, is_staff=0, initial_balance=0.0)

            # Record transaction (no specific product)
            record_transaction(
                product_id=None,
                quantity=None,
                total_amount=CARD_PRICE,
                payment_method="card_purchase",
                rfid_user_id=user_id,
            )

            messagebox.showinfo(
                "Card Issued",
                f"New RFID card created.\nCard ID (simulate UID): {uid}"
            )
            self.build_main_menu()

        tk.Button(
            frame,
            text="Confirm and issue card",
            font=("Arial", 14),
            command=confirm_purchase,
            bg="#4CAF50",
            fg=self.current_theme["button_fg"],
        ).pack(pady=15)

        tk.Button(
            frame,
            text="Cancel and go back",
            command=self.build_main_menu,
            bg="#E53935",
            fg=self.current_theme["button_fg"],
        ).pack(pady=5)

        self.add_theme_toggle_footer()

    # ---------- RFID Reload ----------

    def reload_card_flow(self):
        """Customer reloads an existing RFID card balance."""
        # Simulate RFID scan by asking for card ID
        uid = simpledialog.askstring(
            "RFID Reload",
            "Enter RFID Card ID (simulate tap):",
            parent=self,
        )
        if not uid:
            self.build_main_menu()
            return

        user = get_user_by_uid(uid)
        if not user:
            messagebox.showerror("Error", "Card not found. Please buy a new card first.")
            self.build_main_menu()
            return

        self.clear_screen()
        cash_session.reset()

        frame = tk.Frame(self, bg=self.current_theme["bg"])
        frame.pack(expand=True, fill=tk.BOTH)

        tk.Label(frame, text="Reload RFID Card", font=("Arial", 20)).pack(pady=10)
        tk.Label(frame, text=f"Card ID: {uid}", font=("Arial", 12)).pack(pady=2)
        tk.Label(frame, text=f"Current Balance: ₱{user['balance']:.2f}", font=("Arial", 12)).pack(pady=5)

        amount_var = tk.DoubleVar(value=0.0)

        tk.Label(frame, text="Amount to Load:", font=("Arial", 14)).pack(pady=5)
        tk.Label(frame, textvariable=amount_var, font=("Arial", 22)).pack()

        btn_frame = tk.Frame(frame, bg=self.current_theme["bg"])
        btn_frame.pack(pady=15)

        def add_money(val):
            cash_session.add(val)
            amount_var.set(cash_session.get_amount())

        tk.Button(btn_frame, text="+₱1", width=8, command=lambda: add_money(1)).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="+₱5", width=8, command=lambda: add_money(5)).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="+₱10", width=8, command=lambda: add_money(10)).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="+₱20", width=8, command=lambda: add_money(20)).pack(side=tk.LEFT, padx=5)

        def confirm_reload():
            amount = cash_session.get_amount()
            if amount <= 0:
                messagebox.showwarning("No amount", "Please add some amount to reload.")
                return

            new_balance = user["balance"] + amount
            update_user_balance(user["id"], new_balance)

            record_transaction(
                product_id=None,
                quantity=None,
                total_amount=amount,
                payment_method="rfid_reload",
                rfid_user_id=user["id"],
            )

            messagebox.showinfo(
                "Reload Successful",
                f"New balance: ₱{new_balance:.2f}"
            )
            self.build_main_menu()

        tk.Button(
            frame,
            text="Add balance",
            font=("Arial", 14),
            command=confirm_reload,
            bg="#4CAF50",
            fg=self.current_theme["button_fg"],
        ).pack(pady=15)

        tk.Button(
            frame,
            text="Cancel and go back",
            command=self.build_main_menu,
            bg="#E53935",
            fg=self.current_theme["button_fg"],
        ).pack(pady=5)

        self.add_theme_toggle_footer()

        self.apply_theme_to_widget(self)

    # ---------- RFID Purchase Payment ----------

    def rfid_payment_flow(self, total_amount: float):
        """Purchase products using RFID card balance."""
        uid = simpledialog.askstring(
            "RFID Payment",
            "Enter RFID Card ID (simulate tap):",
            parent=self,
        )
        if not uid:
            self.build_main_menu()
            return

        user = get_user_by_uid(uid)
        if not user:
            messagebox.showerror("Error", "Card not found.")
            self.build_main_menu()
            return

        if user["balance"] < total_amount:
            messagebox.showerror("Error", "Insufficient card balance.")
            self.build_main_menu()
            return

        new_balance = user["balance"] - total_amount
        update_user_balance(user["id"], new_balance)

        p = self.current_product
        q = self.current_quantity

        self.show_wait_screen("Processing RFID payment and dispensing...")
        try:
            decrement_stock(p["id"], q)
            dispense_from_slot(p["slot_number"], q)
            record_transaction(
                product_id=p["id"],
                quantity=q,
                total_amount=total_amount,
                payment_method="rfid_purchase",
                rfid_user_id=user["id"],
            )
            messagebox.showinfo(
                "Thank you",
                f"Payment successful.\nRemaining balance: ₱{new_balance:.2f}\nPlease take your product."
            )
        except Exception as e:
            messagebox.showerror("Error", str(e))
        finally:
            self.build_main_menu()

    # ---------- Restock (Staff) ----------

    def enter_restock_mode(self):
        """Authenticate staff using RFID card ID."""
        uid = simpledialog.askstring(
            "Staff Login",
            "Enter Staff RFID Card ID (simulate tap):",
            parent=self,
        )
        if not uid:
            return

        user = get_user_by_uid(uid)
        if not user or not user["is_staff"]:
            messagebox.showerror("Access Denied", "Invalid staff card.")
            return

        self.show_restock_screen(user)

    # ---------- Admin Dashboard ----------

    def enter_admin_dashboard(self):
        """Authenticate admin using username/password, then show dashboard."""
        stored_username, stored_hash = get_admin_credentials()
        if not stored_username or not stored_hash:
            messagebox.showerror("Error", "Admin credentials not configured.")
            return

        username = simpledialog.askstring(
            "Admin Login",
            "Enter admin username:",
            parent=self,
        )
        if username is None:
            return

        password = simpledialog.askstring(
            "Admin Login",
            "Enter admin password:",
            parent=self,
            show="*",
        )
        if password is None:
            return

        pwd_hash = hashlib.sha256(password.encode("utf-8")).hexdigest()
        if username != stored_username or pwd_hash != stored_hash:
            messagebox.showerror("Access Denied", "Invalid admin credentials.")
            return

        admin_user = {"name": username, "rfid_uid": ""}
        self.show_admin_dashboard(admin_user)

    def show_admin_dashboard(self, staff_user):
        self.clear_screen()

        stats = get_admin_overview_stats()

        # Layout: left sidebar + main content like a dashboard
        container = tk.Frame(self, bg=self.current_theme["bg"])
        container.pack(fill=tk.BOTH, expand=True)

        sidebar = tk.Frame(container, bg=self.current_theme["bg"], width=180)
        sidebar.pack(side=tk.LEFT, fill=tk.Y)

        main = tk.Frame(container, bg=self.current_theme["bg"])
        main.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        # Sidebar navigation
        tk.Label(
            sidebar,
            text="Admin",
            font=("Arial", 16, "bold"),
            bg=self.current_theme["bg"],
            fg=self.current_theme["fg"],
        ).pack(pady=(15, 10))

        tk.Button(
            sidebar,
            text="Overview",
            font=("Arial", 12),
            anchor="w",
            width=18,
            command=lambda: self.show_admin_dashboard(staff_user),
        ).pack(pady=2, padx=10)
        tk.Button(
            sidebar,
            text="Sales Reports",
            font=("Arial", 12),
            anchor="w",
            width=18,
            command=self.show_sales_reports_screen,
        ).pack(pady=2, padx=10)
        tk.Button(
            sidebar,
            text="Change Credentials",
            font=("Arial", 12),
            anchor="w",
            width=18,
            command=self.change_admin_credentials_screen,
        ).pack(pady=2, padx=10)
        tk.Button(
            sidebar,
            text="Back to Main",
            font=("Arial", 12),
            anchor="w",
            width=18,
            command=self.build_main_menu,
        ).pack(pady=(20, 0), padx=10)

        # Main header
        header = tk.Frame(main, bg=self.current_theme["bg"])
        header.pack(fill=tk.X, pady=10, padx=20)

        tk.Label(
            header,
            text="Overview",
            font=("Arial", 18, "bold"),
            bg=self.current_theme["bg"],
            fg=self.current_theme["fg"],
        ).pack(side=tk.LEFT)

        tk.Label(
            header,
            text=f"Admin: {staff_user['name'] or staff_user['rfid_uid']}",
            font=("Arial", 10),
            bg=self.current_theme["bg"],
            fg=self.current_theme["fg"],
        ).pack(side=tk.RIGHT)

        # Metric cards row
        metrics_frame = tk.Frame(main, bg=self.current_theme["bg"])
        metrics_frame.pack(fill=tk.X, padx=20, pady=10)

        def metric_card(parent, title, value_text):
            card = tk.Frame(parent, bg=self.current_theme["button_bg"], bd=1, relief="solid")
            card.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=5)
            tk.Label(
                card,
                text=title,
                font=("Arial", 10),
                bg=self.current_theme["button_bg"],
                fg=self.current_theme["button_fg"],
            ).pack(anchor="w", padx=8, pady=(6, 0))
            tk.Label(
                card,
                text=value_text,
                font=("Arial", 16, "bold"),
                bg=self.current_theme["button_bg"],
                fg=self.current_theme["button_fg"],
            ).pack(anchor="w", padx=8, pady=(0, 8))

        metric_card(metrics_frame, "Total Sales", f"₱{stats['total_sales']:.2f}")
        metric_card(metrics_frame, "Orders", str(stats["orders"]))
        metric_card(metrics_frame, "Active Customers", str(stats["active_customers"]))
        metric_card(metrics_frame, "Low-stock Products", str(stats["low_stock"]))

        # Placeholder area for future charts / tables
        body = tk.Frame(main, bg=self.current_theme["bg"])
        body.pack(fill=tk.BOTH, expand=True, padx=20, pady=(5, 20))

        tk.Label(
            body,
            text="Use the sidebar to open detailed sales reports\nor manage admin credentials.",
            font=("Arial", 11),
            bg=self.current_theme["bg"],
            fg=self.current_theme["fg"],
            justify="left",
        ).pack(anchor="nw", pady=10)

        self.apply_theme_to_widget(self)

    def show_restock_screen(self, staff_user):
        self.clear_screen()

        tk.Label(self, text=f"Restock Mode - Staff: {staff_user['name'] or staff_user['rfid_uid']}",
                 font=("Arial", 16)).pack(pady=10)

        products = get_all_products()
        list_frame = tk.Frame(self, bg=self.current_theme["bg"])
        list_frame.pack(expand=True, fill=tk.BOTH)

        for idx, p in enumerate(products):
            row = tk.Frame(list_frame, bg=self.current_theme["bg"])
            row.pack(fill=tk.X, padx=10, pady=5)

            info = f"Slot {p['slot_number']} - {p['name']} | {p['current_stock']}/{p['capacity']} | ₱{p['price']:.2f}"
            tk.Label(row, text=info, anchor="w").pack(side=tk.LEFT, expand=True)

            tk.Button(
                row,
                text="Restock",
                command=lambda prod=p: self.restock_product_dialog(prod)
            ).pack(side=tk.RIGHT, padx=5)

        tk.Button(self, text="Exit Restock Mode", command=self.build_main_menu).pack(pady=10)

        self.add_theme_toggle_footer()

    def restock_product_dialog(self, product_row):
        max_add = product_row["capacity"] - product_row["current_stock"]
        if max_add <= 0:
            messagebox.showinfo("Full", "This tray is already full.")
            return

        amount = simpledialog.askinteger(
            "Restock Amount",
            f"How many items to add? (max {max_add})",
            minvalue=1,
            maxvalue=max_add,
            parent=self,
        )
        if amount is None:
            return

        new_price = simpledialog.askfloat(
            "New Price",
            f"Enter new price per piece (current ₱{product_row['price']:.2f}), or cancel to keep:",
            parent=self,
        )

        try:
            restock_product(product_row["id"], amount, new_price)
            messagebox.showinfo(
                "Restocked",
                f"Added {amount} items to {product_row['name']}."
            )
        except Exception as e:
            messagebox.showerror("Error", str(e))

        # Refresh restock screen
        # Fetch a dummy staff user (not needed for display name here)
        dummy_staff = {"name": "", "rfid_uid": ""}
        self.show_restock_screen(dummy_staff)

    # ---------- Admin: Sales Reports ----------

    def show_sales_reports_screen(self):
        """Show a list of generated sales report Excel files for admin to open."""
        self.clear_screen()

        tk.Label(
            self,
            text="Sales Reports",
            font=("Arial", 18),
        ).pack(pady=10)

        reports = list_sales_reports()
        list_frame = tk.Frame(self, bg=self.current_theme["bg"])
        list_frame.pack(expand=True, fill=tk.BOTH, padx=20, pady=10)

        if not reports:
            tk.Label(
                list_frame,
                text="No sales reports found. Generate one first.",
                font=("Arial", 12),
            ).pack(pady=10)
        else:
            for idx, report_path in enumerate(reports):
                row = tk.Frame(list_frame, bg=self.current_theme["bg"])
                row.pack(fill=tk.X, pady=5)

                label_text = f"{report_path.name}"
                tk.Label(row, text=label_text, anchor="w").pack(side=tk.LEFT, expand=True)

                def make_open_cmd(p=report_path):
                    def _cmd():
                        try:
                            open_sales_report(p)
                        except Exception as exc:
                            messagebox.showerror("Open Failed", str(exc))
                    return _cmd

                tk.Button(
                    row,
                    text="Open",
                    command=make_open_cmd(),
                ).pack(side=tk.RIGHT, padx=5)

        tk.Button(
            self,
            text="Back to Admin Dashboard",
            font=("Arial", 12),
            command=lambda: self.enter_admin_dashboard(),
        ).pack(pady=10)

        self.add_theme_toggle_footer()

    # ---------- Admin: Change Credentials ----------

    def change_admin_credentials_screen(self):
        """Allow the logged-in admin to change username and password."""
        new_username = simpledialog.askstring(
            "Change Admin Username",
            "Enter new admin username:",
            parent=self,
        )
        if not new_username:
            return

        new_password = simpledialog.askstring(
            "Change Admin Password",
            "Enter new admin password:",
            parent=self,
            show="*",
        )
        if not new_password:
            return

        confirm_password = simpledialog.askstring(
            "Confirm Admin Password",
            "Re-enter new admin password:",
            parent=self,
            show="*",
        )
        if new_password != confirm_password:
            messagebox.showerror("Error", "Passwords do not match.")
            return

        try:
            update_admin_credentials(new_username, new_password)
            messagebox.showinfo(
                "Success",
                "Admin username and password have been updated.",
            )
        except Exception as exc:
            messagebox.showerror("Error", f"Failed to update admin credentials: {exc}")

    # ---------- Sales Report Export ----------

    def export_sales_report_ui(self):
        """Generate an Excel file with transaction, daily, and monthly sales."""
        try:
            path = export_sales_report()
            messagebox.showinfo(
                "Export Complete",
                f"Sales report saved as:\n{path}"
            )
        except Exception as e:
            messagebox.showerror("Export Failed", str(e))


# ======================
#  MAIN ENTRYPOINT
# ======================

def main():
    init_db()
    gpio_init()

    app = MainApp()
    try:
        app.mainloop()
    finally:
        GPIO.cleanup()

if __name__ == "__main__":
    main()